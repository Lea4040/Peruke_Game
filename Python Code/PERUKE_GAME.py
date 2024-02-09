"""-----------import necessary class/Packages/Libraries ------------------"""

from tkinter import *
from openpyxl import *
import random as rand


mainWindow = Tk()




"""-----------color used to design the program----------"""

# #CED3DA: light blue/light gray (background)
# #d12929: red
# #c8d0d9: dark red
# #6aa84f: green




"""-----------define the frames in tkinter------------------"""

# Game Access: Register / Log in Sheet
game_sheet = load_workbook(filename="Peruke_game_records.xlsx")

# Frame 1 : Log in & registration
frame1_1 = Frame(mainWindow, bg="#CED3DA")  # Introducing the game
frame1_2 = Frame(mainWindow, bg="#CED3DA")  # Register / Log in
frame1_3 = Frame(mainWindow, bg="#CED3DA")  # players information (Main Menu)

# Frame 2: Main Game (;
frame2 = Frame(mainWindow, bg="#CED3DA")    # the main game

# Frame 3: Result & record the scores
frame3 = Frame(mainWindow, bg="#CED3DA")  # the result





"""-----------assign the tkinter widgets (the one needs to be called) -----------------"""

# Variables:
username_input = Entry(frame1_2, width=30, font="Century 16")
password_input = Entry(frame1_2, show="*", width=30, font="Century 16")
player1_name_input = Entry(frame1_3, width=30, font="Century 16")
player2_name_input = Entry(frame1_3, width=30, font="Century 16")
player1_label = Label(frame2)
player2_label = Label(frame2)


# Registration error message:
error_message1 = Label(frame1_2, text="Username or password is already used!", bg="#CED3DA", font="Century 11")


# Log in error message:
error_message2 = Label(frame1_2, text="Username or Password is Wrong!", bg="#CED3DA", font="Century 11")





"""-----------------------define the discs list/photos/etc.----------------------------"""

# Set Discs Photos:
safe_disc_photo = [PhotoImage(file="1_safe.png"), PhotoImage(file="2_safe.png"), PhotoImage(file="3_safe.png"),
                   PhotoImage(file="4_safe.png"), PhotoImage(file="5_safe.png"), PhotoImage(file="6_safe.png")]

target_disc_photo = [PhotoImage(file="6_target.png"), PhotoImage(file="5_target.png"), PhotoImage(file="4_target.png"),
                     PhotoImage(file="3_target.png"), PhotoImage(file="2_target.png"), PhotoImage(file="1_target.png")]


# Set Primary & secondary Rows for players

# list for all discs
disc_list = []

# Player1:
player1_primaryRow = []
for i in range(len(safe_disc_photo)):
    player1_primaryRow.append(
        Button(frame2, image=safe_disc_photo[i], text=i+1, relief='flat', border=0, bg="#CED3DA",
               activebackground="#CED3DA", activeforeground="#CED3DA"))
    disc_list.append(player1_primaryRow[i])

player1_secondaryRow = []
for j in range(len(target_disc_photo)):
    player1_secondaryRow.append(
        Button(frame2, image=target_disc_photo[j], text=6-j, relief='flat', border=0, bg="#CED3DA",
               activebackground="#CED3DA", activeforeground="#CED3DA"))
    disc_list.append(player1_secondaryRow[j])


# Player2:
player2_primaryRow = []
for m in range(len(safe_disc_photo)):
    player2_primaryRow.append(
        Button(frame2, image=safe_disc_photo[m], text=m+1, relief='flat', border=0, bg="#CED3DA",
               activebackground="#CED3DA", activeforeground="#CED3DA"))
    disc_list.append(player2_primaryRow[m])

player2_secondaryRow = []
for n in range(len(target_disc_photo)):
    player2_secondaryRow.append(
        Button(frame2, image=target_disc_photo[n], text=6-n,relief='flat', border=0, bg="#CED3DA",
               activebackground="#CED3DA", activeforeground="#CED3DA"))
    disc_list.append(player2_secondaryRow[n])





"""-----------define dice list/photos/etc------------------"""

# List of Dice Photos
dice_photos = [PhotoImage(file="Die1.png", name="1"), PhotoImage(file="Die2.png", name="2"), PhotoImage(file="Die3.png", name="3"),
               PhotoImage(file="Die4.png", name="4"), PhotoImage(file="Die5.png", name="5"), PhotoImage(file="Die6.png", name="6")]


# List of Dice Labels (3 Dice) - (add Photos to Label):
dice = []
for i in range(0, 3):
    dice.append(Label(frame2, border=0, image=dice_photos[5], text=6))


# Button to roll the dice:
roll_dice_button = Button(frame2, text="Roll dice", font="Century 20 bold", height=1, width=10, bg="#c8d0d9", activebackground="#c8d0d9", relief='solid')


# a list to store the integer value when the dice is rolled
dice_numbers = []








"""------------Integers for counting score/press buttons/rounds/etc------------------------"""

# variable to count how many time you press the "roll dice button"  (count rounds)
# + to guide the program to which function is available to use      (odd = player1) / (even = player2)
press_dice_button_count = IntVar()
press_dice_button_count.set(0)



# variable to count how many disc(button) press each player does
player1_disc_press = IntVar()
player1_disc_press.set(0)
player2_disc_press = IntVar()
player2_disc_press.set(0)



# variable to count how many move available for each player (depending on the result of rolling the dice)
player1_available_moves = IntVar()
player1_available_moves.set(0)
player2_available_moves = IntVar()
player2_available_moves.set(0)



# variable to count the score for each player
player1_score = IntVar()
player1_score.set(0)
player2_score = IntVar()
player2_score.set(0)






"""-------------------------Functions for frame1_1 to frame1_3----------------------------"""

# functions to clear each frame:
def clearFrame1_1():
    frame1_1.grid_remove()
def clearFrame1_2():
    frame1_2.grid_remove()
def clearFrame1_3():
    frame1_3.grid_remove()
def clearFrame2():
    frame2.grid_remove()
def clearFrame3():
    frame3.grid_remove()



# checking database in excel sheet (True = username or password exist) / (False = username or password DOESN'T exist)
def check_username_and_password():
    sheet = game_sheet.active
    for row in range(1, sheet.max_row + 1):
        if sheet[f"A{row}"].value == username_input.get() or sheet[f"B{row}"].value == password_input.get():
            return True

    return False



# (Register): Check the length of Entry (Empty Entry NOT ACCEPTED)
def check_entry_for_registration():
    if len(username_input.get()) == 0 or len(password_input.get()) == 0:
        pass
    else:
        username_and_password_registration()


# (Register): Adding username and password to the excel sheet
def username_and_password_registration():
    if check_username_and_password():
        error_message1.grid(pady=10, row=3, column=0, columnspan=2)
    else:
        sheet = game_sheet.active
        sheet[f"A{sheet.max_row + 1}"] = username_input.get()
        sheet[f"B{sheet.max_row}"] = password_input.get()
        sheet[f"C{sheet.max_row}"] = 0
        sheet[f"D{sheet.max_row}"] = 0
        sheet[f"E{sheet.max_row}"] = 0
        sheet[f"F{sheet.max_row}"] = 0
        game_sheet.save(filename="Peruke_game_records.xlsx")
        Frame1_3()



# (Log in): Check the length of Entry (Empty Entry NOT ACCEPTED)
def check_entry_for_login():
    if len(username_input.get()) == 0 or len(password_input.get()) == 0:
        pass
    else:
        username_and_password_login()


# (Log in): Check if the Entered username or password is in the excel sheet
def username_and_password_login():
    if not check_username_and_password():
        error_message2.grid(pady=10, row=3, column=0, columnspan=2)
    else:
        Frame1_3()




# Frame1_2 (1):   Register Function:
def registration_frame():
    clearFrame1_1()
    username_label = Label(frame1_2, text="Enter your username: ", bg="#CED3DA", font="Century 20").grid(padx=30, pady=100, row=0, column=0)
    username_input.grid(padx=150, pady=100, row=0, column=1)
    password_label = Label(frame1_2, text="Enter your password: ", bg="#CED3DA", font="Century 20").grid(padx=30, pady=100, row=1, column=0)
    password_input.grid(padx=150, pady=100, row=1, column=1)
    register_button = Button(frame1_2, text="Register", font="Century 20 bold", height=1, width=10, bg="#c8d0d9", activebackground="#c8d0d9", relief='solid',  command=check_entry_for_registration).grid(pady=100, row=2, column=0, columnspan=2)
    frame1_2.grid()



# Frame1_2 (2):   Log in Function:
def login_frame():
    clearFrame1_1()
    username_label = Label(frame1_2, text="Enter your username: ", bg="#CED3DA", font="Century 20").grid(padx=30, pady=100, row=0, column=0)
    username_input.grid(padx=150, pady=100, row=0, column=1)
    password_label = Label(frame1_2, text="Enter your password: ", bg="#CED3DA", font="Century 20").grid(padx=30, pady=100, row=1, column=0)
    password_input.grid(padx=150, pady=100, row=1, column=1)
    login_button = Button(frame1_2, text="Log in", font="Century 20 bold", height=1, width=10, bg="#c8d0d9", activebackground="#c8d0d9", relief='solid', command=check_entry_for_login).grid(pady=100, row=2, column=0, columnspan=2)
    frame1_2.grid()



# Frame 1_3:    Main Menu of The Game
def Frame1_3():
    clearFrame1_2()
    welcoming_message = Label(frame1_3, text="Welcome to peruke game").grid()
    player1_name_label = Label(frame1_3, text="Player 1 Name: ", bg="#CED3DA", font="Century 20").grid(padx=30, pady=100, row=0, column=0)
    player1_name_input.grid(padx=200, pady=100, row=0, column=1)
    player2_name_label = Label(frame1_3, text="Player 2 Name: ", bg="#CED3DA", font="Century 20").grid(padx=30, pady=100, row=1, column=0)
    player2_name_input.grid(padx=200, pady=100, row=1, column=1)
    game_start = Button(frame1_3, text="Start",  font="Century 20 bold", height=1, width=10, bg="#c8d0d9", activebackground="#c8d0d9", relief='solid', command=check_entry_for_players_names).grid(pady=100, row=2, column=0, columnspan=2)
    frame1_3.grid()



# Check the length of the players' name entry (Empty Entry NOT ACCEPTED)
def check_entry_for_players_names():
    if len(player1_name_input.get()) == 0 or len(player2_name_input.get()) == 0:
        pass
    else:
        Frame2()





# Frame2: Main Game
def Frame2():
    clearFrame1_3()

    # player 1 name
    player1_label.configure(text=player1_name_input.get(), font="Century 16", background='#6aa84f', width=75, height=2)
    player1_label.grid(row=0, column=0, columnspan=6, padx=5, pady=7)
    # player 1 primary row
    for i in range(len(player1_primaryRow)):
        player1_primaryRow[i].grid(row=1, column=i, padx=10, pady=5)
    # player 1 secondary row
    for j in range(len(player1_secondaryRow)):
        player1_secondaryRow[j].grid(row=2, column=j, padx=10, pady=5)



    # dice in the main Window
    roll_dice_button.configure(command=roll_dice)
    roll_dice_button.grid(row=3, column=0, columnspan=6, padx=5, pady=7)

    dice[0].grid(row=4, column=0, columnspan=2, padx=2, pady=5)
    dice[1].grid(row=4, column=2, columnspan=2, padx=2, pady=5)
    dice[2].grid(row=4, column=4, columnspan=2, padx=2, pady=5)



    # player 2 primary row
    for m in range(len(player2_secondaryRow)):
        player2_secondaryRow[m].grid(row=5, column=m, padx=10, pady=5)
    # player 2 secondary row
    for n in range(len(player2_primaryRow)):
        player2_primaryRow[n].grid(row=6, column=n, padx=10, pady=5)
    # player 2 name
    player2_label.configure(text=player2_name_input.get(), font="Century 16", background='#d12929', width=75, height=2)
    player2_label.grid(row=7, column=0, columnspan=6, padx=5, pady=7)


    frame2.grid()





# Rolling the dice & changing the photo of a die & saving the integer value of the dice in dice_numbers
def roll_dice():
    # if roll dice button is pressed (the beginning of the game)
    press_dice_button_count.set(press_dice_button_count.get() + 1)

    # get random value for the dice
    for i in range(len(dice)):
        random = rand.choice(dice_photos)
        dice[i].configure(image=random, text=f"{random.name}")
        dice[i].grid(row=4, column=0 + (2*i), columnspan=2, padx=2, pady=5)
        dice_numbers.append(int(dice[i]['text']))



    # an algorithm used to delete extra dice in the dice_numbers list
    d = []
    for n in range(len(dice_numbers)):
        count_disabled_disc = 0
        for m in range(len(disc_list)):
            if disc_list[m]['state'] == DISABLED and disc_list[m]['text'] == dice_numbers[n]:
                count_disabled_disc += 1
            if count_disabled_disc == 4:
                d.append(dice_numbers[n])
                break
    for k in range(len(d)):
        dice_numbers.remove(d[k])



    # check who is turn now: (odd = player 1)   (even = player 2)
    roll_dice_button.configure(state=DISABLED)      # the button will be enabled in the beginning of others players turn

    if press_dice_button_count.get() == 1:
        round1_for_player1()
    elif press_dice_button_count.get() == 2:
        round1_for_player2()
    elif press_dice_button_count.get() % 2 == 1:
        disc_action_for_player1()    # player 1 turn's function
    else:
        disc_action_for_player2()    # player 2 turn's function





# Round 1 for player 1 (only making their secondary row safe):
def round1_for_player1():

    # all the discs will have action (do nothing) by default ... unless the disc matches the dice number
    for d in range(len(disc_list)):
        disc_list[d].configure(command=do_nothing)

    # check which disc will be changing the command
    for i in range(0, 3):
        for j in range(0, 6):
            if dice[i]['image'] == dice_photos[j].name:
                player1_secondaryRow[5 - j].configure(state=DISABLED, bg='green', command=lambda disc=player1_secondaryRow[5 - j]: make_disc_safe(disc))

    # check how many moves available for player 1
    for i in range(0, 6):
        if player1_secondaryRow[i]['state'] == DISABLED:
            player1_secondaryRow[i].configure(state=NORMAL)
            player1_available_moves.set(player1_available_moves.get() + 1)




# Round 1 for player 2 (only making their secondary row safe):
def round1_for_player2():

    # check which disc will be changing the command
    for i in range(0, 3):
        for j in range(0, 6):
            if dice[i]['image'] == dice_photos[j].name:
                player2_secondaryRow[5 - j].configure(state=DISABLED, bg='green', command=lambda disc=player2_secondaryRow[5 - j]: make_disc_safe(disc))

    # check how many moves available for player 2
    for i in range(0, 6):
        if player2_secondaryRow[i]['state'] == DISABLED:
            player2_secondaryRow[i].configure(state=NORMAL)
            player2_available_moves.set(player2_available_moves.get() + 1)






# function for player 1: make their rows safe and take other players row
def disc_action_for_player1():

    for i in range(0, 3):   # dice range

        for j in range(0, 6):
            # make primary row safe for player 1
            if player1_primaryRow[j]['state'] == DISABLED:
                pass
            elif (dice[i]['image'] == dice_photos[j].name) and (player1_primaryRow[j]['image'] == target_disc_photo[5 - j].name):
                player1_primaryRow[j].configure(underline=0, bg="green", command=lambda disc=player1_primaryRow[j]: make_disc_safe(disc))


            # make secondary row safe for player 1
            if player1_secondaryRow[5 - j]['state'] == DISABLED:
                pass
            elif (dice[i]['image'] == dice_photos[j].name) and (player1_secondaryRow[5 - j]['image'] == target_disc_photo[5 - j].name):
                player1_secondaryRow[5 - j].configure(underline=0, bg="green", command=lambda disc=player1_secondaryRow[5 - j]: make_disc_safe(disc))



            # make secondary row target for player 2
            if player2_secondaryRow[5 - j]['state'] == DISABLED:
                pass
            elif dice[i]['image'] == dice_photos[j].name:
                player2_secondaryRow[5 - j].configure(underline=0, bg="green", command=lambda disc=player2_secondaryRow[5 - j]: make_disc_target(disc))



            # make primary row target for player 2
            if player2_primaryRow[j]['state'] == DISABLED:
                pass
            elif dice[i]['image'] == dice_photos[j].name:
                player2_primaryRow[j].configure(underline=0, bg="green", command=lambda disc=player2_primaryRow[j]: make_disc_target(disc))


    # check how many moves available for player 1
    for d in disc_list:
        if d['underline'] == 0:
            d.configure(underline=-1)
            player1_available_moves.set(player1_available_moves.get() + 1)



    # eliminate extra dice (if there is no dice matches any disc)
    d = []
    for r in range(len(dice_numbers)):
        count = 0
        for s in range(len(disc_list)):
            if (dice_numbers[r] == disc_list[s]['text']) and (disc_list[s]['bg'] == "#CED3DA"):
                count += 1
            if count == 4:
                d.append(dice_numbers[r])
                break
    for k in range(len(d)):
        dice_numbers.remove(d[k])



    check_whose_turn()







# function for player 2: make their rows safe and take other players row
def disc_action_for_player2():


    for i in range(0, 3):  # dice range

        for j in range(0, 6):
            # make primary row safe for player 2
            if player2_primaryRow[j]['state'] == DISABLED:
                pass
            elif (dice[i]['image'] == dice_photos[j].name) and (player2_primaryRow[j]['image'] == target_disc_photo[5 - j].name):
                player2_primaryRow[j].configure(underline=0, bg="green", command=lambda disc=player2_primaryRow[j]: make_disc_safe(disc))



            # make secondary row safe for player 2
            if player2_secondaryRow[5 - j]['state'] == DISABLED:
                pass
            elif (dice[i]['image'] == dice_photos[j].name) and (player2_secondaryRow[5 - j]['image'] == target_disc_photo[5 - j].name):
                player2_secondaryRow[5 - j].configure(underline=0, bg="green", command=lambda disc=player2_secondaryRow[5 - j]: make_disc_safe(disc))




            # make secondary row target for player 1
            if player1_secondaryRow[5 - j]['state'] == DISABLED:
                pass
            elif dice[i]['image'] == dice_photos[j].name:
                player1_secondaryRow[5 - j].configure(underline=0, bg="green", command=lambda disc=player1_secondaryRow[5 - j]: make_disc_target(disc))

            # make primary row target for player 1
            if player1_primaryRow[j]['state'] == DISABLED:
                pass
            elif dice[i]['image'] == dice_photos[j].name:
                player1_primaryRow[j].configure(underline=0, bg="green", command=lambda disc=player1_primaryRow[j]: make_disc_target(disc))




    # check how many moves available for player 2
    for d in disc_list:
        if d['underline'] == 0:
            d.configure(underline=-1)
            player2_available_moves.set(player2_available_moves.get() + 1)



    # eliminate extra dice (if there is no dice matches any disc)
    d = []
    for r in range(len(dice_numbers)):
        count = 0
        for s in range(len(disc_list)):
            if (dice_numbers[r] == disc_list[s]['text']) and (disc_list[s]['bg'] == "#CED3DA"):
                count += 1
            if count == 4:
                d.append(dice_numbers[r])
                break
    for k in range(len(d)):
        dice_numbers.remove(d[k])



    check_whose_turn()









# make disc safe, argument = disc:
def make_disc_safe(disc):

    if disc['text'] in dice_numbers:
        # increase the count button for players
        if press_dice_button_count.get() % 2 == 1:
            player1_disc_press.set(player1_disc_press.get() + 1)
        else:
            player2_disc_press.set(player2_disc_press.get() + 1)


        # make disc safe by switching the photo of the disc from target to safe
        for i in range(0, 6):
            if (disc in player1_primaryRow) or (disc in player2_primaryRow):
                if disc['image'] == target_disc_photo[5 - i].name:
                    disc.configure(image=safe_disc_photo[i], bg='red')
            else:
                if disc['image'] == target_disc_photo[i].name:
                    disc.configure(image=safe_disc_photo[5-i], bg='red')

        dice_numbers.remove(disc['text'])



    # another algorithm to eliminate extra dice ( if no dice matches the discs)
    for i in range(0, 2):
        count_ = 0
        for i in range(len(disc_list)):
            if (disc['text'] in dice_numbers) and (disc_list[i]['text'] == disc['text']) and (disc_list[i]['bg'] == 'green'):
                count_ += 1

        if count_ == 0 and (disc['text'] in dice_numbers):
            dice_numbers.remove(disc['text'])



    check_whose_turn()






# make disc target, argument = disc
def make_disc_target(disc):


    if disc['text'] in dice_numbers:
        # increase the count button for players
        if press_dice_button_count.get() % 2 == 1:
            player1_disc_press.set(player1_disc_press.get() + 1)
        else:
            player2_disc_press.set(player2_disc_press.get() + 1)



        for i in range(0, 6):
            # primary row
            if (disc in player1_primaryRow) or (disc in player2_primaryRow):
                # make disc target
                if disc['image'] == safe_disc_photo[i].name:
                    disc.configure(image=target_disc_photo[5 - i], bg='red')
                    if dice_numbers.count(disc['text']) == 2:
                        disc.configure(bg='green')

                # take off the disc (by making them disabled)
                elif disc['image'] == target_disc_photo[5 - i].name:
                    if press_dice_button_count.get() % 2 == 1:
                        player1_score.set(player1_score.get() + disc['text'])
                    else:
                        player2_score.set(player2_score.get() + disc['text'])
                    disc.configure(state=DISABLED, bg="#CED3DA")

            # secondary row
            else:
                # make disc target
                if disc['image'] == safe_disc_photo[i].name:
                    disc.configure(image=target_disc_photo[5 - i], bg='red')
                    if dice_numbers.count(disc['text']) == 2:
                        disc.configure(bg='green')

                # take off the disc (by making them disabled)
                elif disc['image'] == target_disc_photo[5 - i].name:
                    if press_dice_button_count.get() % 2 == 1:
                        player1_score.set(player1_score.get() + disc['text'])
                    else:
                        player2_score.set(player2_score.get() + disc['text'])
                    disc.configure(state=DISABLED, bg="#CED3DA")

        dice_numbers.remove(disc['text'])




    # another algorithm to eliminate extra dice (if no dice matches the discs)
    for i in range(0, 2):
        count_ = 0
        for i in range(len(disc_list)):
            if (disc['text'] in dice_numbers) and (disc_list[i]['text'] == disc['text']) and (disc_list[i]['bg'] == 'green'):
                count_ += 1
        if count_ == 0 and (disc['text'] in dice_numbers):
            dice_numbers.remove(disc['text'])






    # algorithm to check if the primary row for player is taken (all discs are DISABLED)
    count1 = 0
    count2 = 0
    for i in range(0, 6):
        if player1_primaryRow[i]['state'] == DISABLED:
            count1 += 1
        if player2_primaryRow[i]['state'] == DISABLED:
            count2 += 1
        if count1 == 6 or count2 == 6:
            Frame3()





    check_whose_turn()







# function to reset players press count + their available moves:
def reset():
    for disc in disc_list:
        disc.configure(relief='flat', bg="#CED3DA", command=do_nothing)

    roll_dice_button.configure(state=NORMAL)
    dice_numbers.clear()
    player1_disc_press.set(0)
    player1_available_moves.set(0)
    player2_disc_press.set(0)
    player2_available_moves.set(0)




# function to be used if a player press on one of the disc accidentally / to change the command for discs (buttons)
def do_nothing():
    pass




# function to check who is players turn
def check_whose_turn():
    # player 1 turn
    if press_dice_button_count.get() % 2 == 1:
        if len(dice_numbers) == 0 or player1_available_moves.get() == 0:
            reset()
            player1_label.configure(background='#d12929')
            player2_label.configure(background='#6aa84f')


    # player 2 turn
    else:
        if len(dice_numbers) == 0 or player2_available_moves.get() == 0:
            reset()
            player2_label.configure(background='#d12929')
            player1_label.configure(background='#6aa84f')





"""-------------------------frame 3----------------------------"""
def Frame3():
    clearFrame2()
    frame3.grid()
    Label(frame3, text="END OF THE GAME", bg="#CED3DA", font="Century 22 bold", fg="#d12929").grid(padx=350, pady=150, row=0, column=0, columnspan=2)


    Label(frame3, text="PLAYERS", bg="#CED3DA", font="Century 16").grid(pady=10, row=1, column=0)
    Label(frame3, text="SCORES", bg="#CED3DA", font="Century 16").grid(pady=10, row=1, column=1)


    # player 1
    if press_dice_button_count.get() % 2 == 1:
        for i in range(0, 12):
            # add the available discs to his score
            if disc_list[i]['state'] == NORMAL:
                player1_score.set(player1_score.get() + int(disc_list[i]['text']))

        for j in range(12, 24):
            # add the target discs from other player to his score
            if disc_list[j]['image'] in target_disc_photo:
                player1_score.set(player1_score.get() + int(disc_list[i]['text']))
            else:
                player2_score.set(player2_score.get() + int(disc_list[i]['text']))



    # player 2
    else:
        for i in range(12, 24):
            # add the available discs to his score
            if disc_list[i]['state'] == NORMAL:
                player2_score.set(player2_score.get() + int(disc_list[i]['text']))

        for j in range(0, 12):
            # add the target discs from other player to his score
            if disc_list[j]['image'] in target_disc_photo:
                player2_score.set(player2_score.get() + int(disc_list[i]['text']))
            else:
                player1_score.set(player1_score.get() + int(disc_list[i]['text']))


    # Displaying the result and edit the excel sheet
    theWinner = " "
    sheet = game_sheet.active
    for row in range(1, sheet.max_row + 1):
        if sheet[f"A{row}"].value == username_input.get():
            sheet[f"C{row}"] = int(sheet[f"C{row}"].value) + 1
            if player1_score.get() > player2_score.get():
                theWinner = f"{player1_name_input.get()} Win The Game !"
                sheet[f"D{row}"] = int(sheet[f"D{row}"].value) + 1

            elif player1_score.get() < player2_score.get():
                theWinner = f"{player2_name_input.get()} Win The Game !"
                sheet[f"E{row}"] = int(sheet[f"e{row}"].value) + 1

            else:
                theWinner = " -- DRAW --"


            if player1_score.get() > sheet[f"F{row}"].value:
                sheet[f"F{row}"] = int(player1_score.get())


    # save the changes
    game_sheet.save(filename="Peruke_game_records.xlsx")


    # player 1 name and score
    Label(frame3, text=player1_name_input.get(), bg="#CED3DA", font="Century 16").grid(pady=10, row=2, column=0)
    Label(frame3, text=player1_score.get(), bg="#CED3DA", font="Century 16").grid(pady=10, row=2, column=1)

    # player 2 name and score
    Label(frame3, text=player2_name_input.get(), bg="#CED3DA", font="Century 16").grid(pady=10, row=3, column=0)
    Label(frame3, text=player2_score.get(), bg="#CED3DA", font="Century 16").grid(pady=10, row=3, column=1)

    # display the winner name / or DRAW
    Label(frame3, text=theWinner, bg="#CED3DA", font="Century 16").grid(padx=350, pady=150, row=4, column=0, columnspan=2)





"""-----------------call the game-----------------"""

# Frame 1:
mainWindow.configure(bg="#CED3DA")
mainWindow.title("Peruke")  # display the name of the game
mainWindow.geometry("1000x800")     # window size
peruke_game_image = PhotoImage(file='Peruke_game_logo.png')
Label(frame1_1, image=peruke_game_image, bg="#CED3DA").grid(padx=175, pady=30)

game_Label = Label(frame1_1, text="Welcome To Peruke Game", bg="#CED3DA", font="Century 20 bold", fg="#d12929").grid(padx=300, pady=40)  # Name of he game on the screen
register_Button = Button(frame1_1, text="Register", font="Century 20 bold", height=1, width=10, bg="#c8d0d9", activebackground="#c8d0d9", relief='solid', command=registration_frame).grid(padx=300)  # register Button
logIn_Button = Button(frame1_1, text="Log in", font="Century 20 bold", height=1, width=10, bg="#c8d0d9", activebackground="#c8d0d9", relief='solid', command=login_frame).grid(padx=310, pady=45)  # Log in Button
frame1_1.grid()


mainWindow.mainloop()
